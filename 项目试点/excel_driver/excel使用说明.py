# # # ！/usr/bin/python3
# # -*- coding: utf-8 -*-
# # 当前项目名称：Selenium3自动化测试实战
# # 文件名称： excel使用说明
# # 登录用户名： yanshaoyou
# # 作者： 闫少友
# # 邮件： 2395969839@qq.com
# # 电话：17855503800
# # 创建时间： 2021/12/9  10:33
#
# # remove_sheet('sheet')  # 删除sheet页
# # create_sheet("sheet", 0)  # 创建sheet页，并指定位置
# # sheet_properties.tabColor = "0072BA"  # 更改工作表的背景颜色
# # merge_cells('A1:B2')  # 合并单元格，保留左上角的值
# # Alignment(horizontal='center', vertical='center')  # 居中，设置格式
# # freeze_panes = 'B2'  # 冻结窗口
# #
# # cell = sheet.cell(row=7, column=2)  # 创建单元格
# # cell.value = "=SUM(A1:B6)"  # 写入公式
# # cell.font = cell.font.copy(bold=True)  # 改变字体
# # sheet.add_image(img, 'B2')  # 插入图片
#
# import openpyxl
#
# # 读取操作
# excel = openpyxl.load_workbook('../data/data.xlsx')  # 读取excel文件
# sheets = excel.sheetnames  # 获取所有的sheet名,方法1
# for sheet in excel:
#     print(sheet.title)  # 获取所有的sheet名,方法2
# sheet = excel.active
# a1 = sheet['A1']
# a2 = sheet['A2']
# a3 = sheet['A3']
# print(a1.value, a2.value, a3.value)  # 读取单个单元格值
# cells = sheet['A1':'B7']  # 获取多行单元格内容
# for c1, c2 in cells:
#     print("{0:3} {1:8}".format(c1.value, c2.value))  # {0：3}，显示格式，前面是2个空格，第3个开始写
#
# # 操作excel文件
# from openpyxl import Workbook
# import time
#
# book = Workbook()
# sheet = book.active
# sheet['A1'] = 56  # 使用键输入单元格内容
# sheet['A2'] = 43
# now = time.strftime('%x')
# sheet['A3'] = now
# book.save('../data/1.xlsx')  # 创建保存文件
#
# sheet.cell(row=2, column=2).value = 2  # 使用行和列进行输入单元格内容
# book.save('../data/2.xlsx')
#
# from openpyxl import Workbook
# import time
# book = Workbook()
# sheet = book.active
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )
# for row in rows:
#     sheet.append(row)  # 在数据底部附加值，append
# book.save('../data/3.xlsx')
# from openpyxl import Workbook
#
# book = Workbook()
# sheet = book.active
#
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )
#
# for row in rows:
#     sheet.append(row)
#
# for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):  # 按行迭代，iter_rows：将工作表中的单元格返回为行
#     for cell in row:
#         print(cell.value, end=" ")
#     print()
#
# for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):  # 按列迭代，iter_cols：将工作表中的单元格返回为列
#     for cell in row:
#         print(cell.value, end=" ")
#     print()
#
# book.save('../data/4.xlsx')
#
# import openpyxl
# import statistics as stats
#
# book = openpyxl.load_workbook('../data/5.xlsx', data_only=True)
# sheet = book.active
# rows = sheet.rows
# values = []
# for row in rows:
#     for cell in row:
#         values.append(cell.value)
#
# print("Number of values: {0}".format(len(values)))
# print("Sum of values: {0}".format(sum(values)))
# print("Minimum value: {0}".format(min(values)))
# print("Maximum value: {0}".format(max(values)))
# print("Mean: {0}".format(stats.mean(values)))
# print("Median: {0}".format(stats.median(values)))
# print("Standard deviation: {0}".format(stats.stdev(values)))
# print("Variance: {0}".format(stats.variance(values)))
#
# from openpyxl import Workbook
#
# wb = Workbook()
# sheet = wb.active
# data = [
#     ['Item', 'Colour'],
#     ['pen', 'brown'],
#     ['book', 'black'],
#     ['plate', 'white'],
#     ['chair', 'brown'],
#     ['coin', 'gold'],
#     ['bed', 'brown'],
#     ['notebook', 'white'],
# ]
# for r in data:
#     sheet.append(r)
#
# sheet.auto_filter.ref = 'A1:B8'  # 筛选范围
# sheet.auto_filter.add_filter_column(1, ['brown', 'white'])  # 列筛选操作，先确定第几列，然后筛选内容
# sheet.auto_filter.add_sort_condition('B2:B8')  # 要筛选的列的起始、终止单元格
#
# wb.save('../data/6.xlsx')
#
#
# from openpyxl import Workbook
#
# book = Workbook()
# sheet = book.active
#
# sheet['A3'] = 39
# sheet['B3'] = 19
#
# rows = [
#     (88, 46),
#     (89, 38),
#     (23, 59),
#     (56, 21),
#     (24, 18),
#     (34, 15)
# ]
#
# for row in rows:
#     sheet.append(row)
#
# print(sheet.dimensions)  # 获得那些实际包含数据的单元格,返回起始-终止，单元格名称
# print("Minimum row: {0}".format(sheet.min_row))  # 有数据的最小单元行数值
# print("Maximum row: {0}".format(sheet.max_row))
# print("Minimum column: {0}".format(sheet.min_column))  # 有数据的最小单元列数值
# print("Maximum column: {0}".format(sheet.max_column))
#
# for c1, c2 in sheet[sheet.dimensions]:  # 获取单元格维度
#     print(c1.value, c2.value)
#
# book.save('../data/7.xlsx')
#
#
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
#
# book = Workbook()
# sheet = book.active
#
# sheet.merge_cells('A1:B2')
#
# cell = sheet.cell(row=1, column=1)
# cell.value = 'Sunny day'
# cell.alignment = Alignment(horizontal='center', vertical='center')
#
# book.save('../data/8.xlsx')
#
#
# from openpyxl import Workbook
# from openpyxl.chart import (
#     Reference,
#     Series,
#     BarChart
# )
#
# book = Workbook()
# sheet = book.active
#
# rows = [
#     ("USA", 46),
#     ("China", 38),
#     ("UK", 29),
#     ("Russia", 22),
#     ("South Korea", 13),
#     ("Germany", 11)
# ]
#
# for row in rows:
#     sheet.append(row)
#
# data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=6)
# categs = Reference(sheet, min_col=1, min_row=1, max_row=6)
#
# chart = BarChart()
# chart.add_data(data=data)
# chart.set_categories(categs)
#
# chart.legend = None
# chart.y_axis.majorGridlines = None
# chart.varyColors = True
# chart.title = "Olympic Gold medals in London"
#
# sheet.add_chart(chart, "A8")
#
# book.save("../data/9.xlsx")
#
# import datetime
# from random import choice
# from time import time
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
#
# # # 设置文件 mingc
# # addr = "openpyxl.xlsx"
# # # 打开文件
# # wb = load_workbook(addr)
# # # 创建一张新表
# # ws = wb.create_sheet()
# # # 第一行输入
# # ws.append(['TIME', 'TITLE', 'A-Z'])
# #
# # # 输入内容（500行数据）
# # for i in range(500):
# #     TIME = datetime.datetime.now().strftime("%H:%M:%S")
# #     TITLE = str(time())
# #     A_Z = get_column_letter(choice(range(1, 50)))
# #     ws.append([TIME, TITLE, A_Z])
# #
# # # 获取最大行
# # row_max = ws.max_row
# # # 获取最大列
# # con_max = ws.max_column
# # # 把上面写入内容打印在控制台
# # for j in ws.rows:    # we.rows 获取每一行数据
# #     for n in j:
# #         print(n.value, end="\t")   # n.value 获取单元格的值
# #     print()
# # # 保存，save（必须要写文件名（绝对地址）默认 py 同级目录下，只支持 xlsx 格式）
# # wb.save(addr)


# from openpyxl import load_workbook
# import pandas as pd
# from openpyxl.chart import LineChart, Reference
#
# target_file = pd.Dataframe()
# target_file.to_excel('target.xlsx')
# sheet_names = ['A', 'B', 'C']
# # 创建3个sheet
# for sheet_name in sheet_names:
#     writer = pd.ExcelWriter('target.xlsx', engine='openpyxl')
#     book = load_workbook('target.xlsx')
#     writer.book = book
#     df.to_excel(writer, sheet_name=sheet_name)
#     writer.save()
#
# # 将数据写入指定sheet中
# class_info = [['class1', 'class2', 'class3'], [1, 2, 3], [2, 3, 4], [3, 4, 5], [5, 6, 7]]
# wb = load_workbook('target.xlsx')
# ws = wb['A']
# for r in range(len(class_info)):
#     col = class_info[r]
#     for c in range(len(col)):
#         ws.cell(r + 1, c + 1).value = class_info[r][c]
# wb.save('target.xlsx')


# from openpyxl.chart import LineChart,Reference
#
# wb=load_workbook('target.xlsx')
# ws=wb['A']
# chart=LineChart()
# chart='hello'
# chart.style=12
# chart.title='just for fun'
# chart.x_axis.title='class No'
# chart.y_axis.title='score'
# #添加不连续区间内数据可用多次添加方式，添加连续区间之间选择最大行列号和最小行列号区域内的数据即可
# data1=Reference(ws,min_col=1,min_row=2,max_col=1,max_row=5)
# data2=Reference(ws,min_col=2,min_row=2,max_col=2,max_row=5)
# data3=Reference(ws,min_col=3,min_row=2,max_col=3,max_row=5)
# chart.add_data(data1,titles_from_data=True)
# chart.add_data(data2,titles_from_data=True)
# chart.add_data(data3,titles_from_data=True)
# chart.shape=4
# ws.add_chart(chart,'E1')
# wb.save('target.xlsx')

import openpyxl
import pandas as pd

path = r"C:\Users\john\Desktop\test.xlsx"
df = pd.read_excel(path,index_col=None)
Result = df.head()

excel_writer = pd.ExcelWriter(path, engine='openpyxl')
book = openpyxl.load_workbook(excel_writer.path)
excel_writer.book = book
Result.to_excel(excel_writer=excel_writer, sheet_name= 'test', index=None)
excel_writer.close()

from openpyxl import load_workbook,Workbook
import pandas as pd
path = r'C:\Users\wangyiming\Desktop\现货系统收款导入文件格式.xlsx'
file_path = r'C:\Users\wangyiming\Desktop\表样.xls'
wb = Workbook()
wb.save(path)
book = load_workbook(path)
print(book)
writer = pd.ExcelWriter(path,engine='openpyxl')
result = pd.read_excel(file_path)
#print('样表：'+result)
writer.book = book
print(book)
result.to_excel(writer,sheet_name='Sheet1',index=False)
writer.save()